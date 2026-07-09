import os
import time
import sqlite3
import pickle
import logging
import numpy as np
import pandas as pd
from datetime import datetime
from flask import render_template, Response, request, redirect, url_for, session, jsonify, send_file, send_from_directory
from werkzeug.security import generate_password_hash, check_password_hash

from database import get_db_connection
import camera

def get_attendance_report_data(emp_search="", dept_filter="", start_date="", end_date=""):
    """
    Computes a list of dicts containing attendance status (Present or Absent) 
    for all employees over the dates present in the attendance records or queried range.
    """
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # 1. Fetch filtered employees
    emp_query = "SELECT EmployeeID, Name, Department, Designation, Phone, Email FROM employees WHERE 1=1"
    emp_params = []
    if emp_search:
        emp_query += " AND (EmployeeID LIKE ? OR Name LIKE ?)"
        emp_params.extend([f"%{emp_search}%", f"%{emp_search}%"])
    if dept_filter:
        emp_query += " AND Department = ?"
        emp_params.append(dept_filter)
        
    cursor.execute(emp_query, emp_params)
    employees = [dict(r) for r in cursor.fetchall()]
    
    # 2. Fetch all distinct dates from the attendance table
    cursor.execute("SELECT DISTINCT Date FROM attendance")
    all_dates_str = [r['Date'] for r in cursor.fetchall()]
    
    # 3. Filter dates by range
    active_dates = set()
    for date_str in all_dates_str:
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            if start_date:
                start_obj = datetime.strptime(start_date, "%Y-%m-%d")
                if date_obj < start_obj:
                    continue
            if end_date:
                end_obj = datetime.strptime(end_date, "%Y-%m-%d")
                if date_obj > end_obj:
                    continue
            active_dates.add(date_obj)
        except Exception as e:
            logging.error(f"Error parsing date {date_str}: {str(e)}")
            
    # Include the specific queried date even if no attendance has been recorded yet (helps visibility)
    if start_date and end_date and start_date == end_date:
        try:
            single_date_obj = datetime.strptime(start_date, "%Y-%m-%d")
            active_dates.add(single_date_obj)
        except ValueError:
            pass
            
    sorted_dates = sorted(list(active_dates), reverse=True)
    if not sorted_dates:
        # Fallback to today if database and filters yield no dates
        sorted_dates = [datetime.today()]
        
    # 4. Fetch all attendance logs to check who is present
    cursor.execute("SELECT EmployeeID, Date, Time, Status FROM attendance")
    attendance_records = cursor.fetchall()
    conn.close()
    
    attendance_lookup = {}
    for r in attendance_records:
        attendance_lookup[(r['EmployeeID'], r['Date'])] = (r['Time'], r['Status'])
        
    # 5. Build full grid of present/absent logs
    results = []
    for date_obj in sorted_dates:
        date_str = date_obj.strftime("%d-%m-%Y")
        for emp in employees:
            emp_id = emp['EmployeeID']
            emp_name = emp['Name']
            dept = emp['Department'] if emp['Department'] else 'N/A'
            phone = emp['Phone'] if emp['Phone'] else 'N/A'
            email = emp['Email'] if emp['Email'] else 'N/A'
            
            key = (emp_id, date_str)
            if key in attendance_lookup:
                time_str, status = attendance_lookup[key]
            else:
                time_str = "N/A"
                status = "Absent"
                
            results.append({
                'attendance_id': None,
                'employee_id': emp_id,
                'name': emp_name,
                'department': dept,
                'designation': emp.get('Designation') or 'N/A',
                'phone': phone,
                'email': email,
                'date': date_str,
                'time': time_str,
                'status': status
            })
            
    # Sort: Date desc, Time desc (making Present logs rise to the top of each date)
    results.sort(
        key=lambda x: datetime.strptime(
            f"{x['date']} {x['time'] if x['time'] != 'N/A' else '00:00:00'}", 
            "%d-%m-%Y %H:%M:%S"
        ), 
        reverse=True
    )
    return results

def register_routes(app):
    @app.before_request
    def check_auth():
        # Require login for all routes except static assets, login page, and health checks
        allowed_routes = ['login', 'static']
        if not session.get('logged_in') and request.endpoint not in allowed_routes and request.endpoint is not None:
            return redirect(url_for('login'))

    @app.route('/video_feed')
    def video_feed():
        if not session.get('logged_in'):
            return Response("Unauthorized", status=401)
        return Response(camera.gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if session.get('logged_in'):
            return redirect(url_for('dashboard'))
            
        error = None
        if request.method == 'POST':
            email = request.form.get('email', '').strip()
            password = request.form.get('password')
            
            # Verify administrator credentials in local SQLite DB
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT Name, PasswordHash FROM admins WHERE Email = ?", (email,))
            admin = cursor.fetchone()
            conn.close()
            
            if admin and check_password_hash(admin['PasswordHash'], password):
                session['logged_in'] = True
                session['username'] = admin['Name']
                logging.info(f"Admin user logged in successfully: {email}")
                return redirect(url_for('dashboard'))
            else:
                error = "Invalid email address or password."
                logging.warning(f"Failed login attempt for email: {email}")
                
        return render_template('login.html', error=error)

    @app.route('/logout')
    def logout():
        session.clear()
        camera.system_mode = 'passive'
        logging.info("Admin logged out.")
        return redirect(url_for('login'))

    @app.route('/')
    def dashboard():
        # Reset system mode to passive when on dashboard
        camera.system_mode = 'passive'
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Statistics
        cursor.execute("SELECT COUNT(*) FROM employees")
        total_employees = cursor.fetchone()[0]
        
        today_date = datetime.now().strftime("%d-%m-%Y")
        cursor.execute("SELECT COUNT(DISTINCT EmployeeID) FROM attendance WHERE Date = ?", (today_date,))
        present_today = cursor.fetchone()[0]
        
        absent_today = max(0, total_employees - present_today)
        conn.close()
        
        return render_template('index.html', 
                               total_employees=total_employees, 
                               present_today=present_today, 
                               absent_today=absent_today,
                               today_date=datetime.now().strftime("%B %d, %Y"))

    @app.route('/register')
    def register_page():
        camera.system_mode = 'passive'
        return render_template('register.html')

    @app.route('/attendance')
    def attendance_page():
        camera.system_mode = 'attendance'
        return render_template('attendance.html')

    @app.route('/employees')
    def employees_page():
        camera.system_mode = 'passive'
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT EmployeeID, Name, Department, Designation, Phone, Email, DateOfJoining, CreatedAt FROM employees")
        employees = cursor.fetchall()
        conn.close()
        
        return render_template('employees.html', employees=employees)

    @app.route('/reports')
    def reports_page():
        camera.system_mode = 'passive'
        return render_template('reports.html')

    # ----------------- APIs & Ajax Operations -----------------

    @app.route('/api/latest_verification')
    def api_latest_verification():
        """Returns the latest checked-in employee details for client alert popups."""
        # Expiry of verification box (expires after 4 seconds)
        if time.time() - camera.latest_match['timestamp'] < 4.0:
            return jsonify(camera.latest_match)
        return jsonify({'employee_id': None})

    @app.route('/api/stats')
    def api_stats():
        """Returns real-time dashboard counts."""
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM employees")
        total = cursor.fetchone()[0]
        
        today_date = datetime.now().strftime("%d-%m-%Y")
        cursor.execute("SELECT COUNT(DISTINCT EmployeeID) FROM attendance WHERE Date = ?", (today_date,))
        present = cursor.fetchone()[0]
        
        absent = max(0, total - present)
        conn.close()
        
        return jsonify({
            'total_employees': total,
            'present_today': present,
            'absent_today': absent
        })

    @app.route('/api/register/start', methods=['POST'])
    def api_register_start():
        """Initiates camera capture and registers employee metadata."""
        emp_id = request.form.get('employee_id', '').strip()
        name = request.form.get('name', '').strip()
        dept = request.form.get('department', '').strip()
        designation = request.form.get('designation', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        doj = request.form.get('date_of_joining', '').strip()
        
        if not emp_id or not name:
            return jsonify({'success': False, 'message': 'Employee ID and Name are required.'})
            
        # Check if ID already exists
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT Name FROM employees WHERE EmployeeID = ?", (emp_id,))
        exists = cursor.fetchone()
        conn.close()
        
        if exists:
            return jsonify({'success': False, 'message': f'Employee ID {emp_id} is already registered to {exists["Name"]}.'})
            
        # Reset capture state
        camera.register_data = {
            'employee_id': emp_id,
            'name': name,
            'department': dept,
            'designation': designation,
            'phone': phone,
            'email': email,
            'date_of_joining': doj,
            'captured_frames': []
        }
        camera.register_status = {
            'count': 0,
            'status': 'capturing',
            'message': 'Starting camera capture...'
        }
        
        camera.system_mode = 'register'
        camera.register_is_capturing = True
        
        # Ensure camera starts
        camera.camera.start()
        
        logging.info(f"Started face capture process for Employee: {name} ({emp_id})")
        return jsonify({'success': True, 'message': 'Face scanner activated.'})

    @app.route('/api/register/status')
    def api_register_status():
        """Polls capture frames status and triggers saving when count hits 30."""
        if camera.register_status['status'] == 'completed':
            # Trigger background processing & saving
            success = camera.save_registered_employee()
            if success:
                camera.register_status['status'] = 'idle'
                camera.system_mode = 'passive'
                return jsonify({'status': 'done', 'message': 'Employee Registered successfully!'})
            else:
                camera.register_status['status'] = 'error'
                camera.register_status['message'] = 'Failed to extract face features. Please scan again.'
                camera.system_mode = 'passive'
                return jsonify({'status': 'failed', 'message': camera.register_status['message']})
                
        return jsonify({
            'status': camera.register_status['status'],
            'count': camera.register_status['count'],
            'message': camera.register_status['message']
        })

    @app.route('/api/employee/delete', methods=['POST'])
    def api_delete_employee():
        """Deletes employee record, face encodings, and local image dataset folder."""
        emp_id = request.form.get('employee_id')
        if not emp_id:
            return jsonify({'success': False, 'message': 'Employee ID is required.'})
            
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            # Get image folder path
            cursor.execute("SELECT ImageFolder FROM employees WHERE EmployeeID = ?", (emp_id,))
            row = cursor.fetchone()
            
            # Delete from Database
            cursor.execute("DELETE FROM employees WHERE EmployeeID = ?", (emp_id,))
            cursor.execute("DELETE FROM attendance WHERE EmployeeID = ?", (emp_id,))
            conn.commit()
            conn.close()
            
            # Delete encodings from FaceHandler cache
            camera.face_handler.remove_employee_face(emp_id)
            
            # Delete local dataset directory
            if row and row['ImageFolder'] and os.path.exists(row['ImageFolder']):
                for filename in os.listdir(row['ImageFolder']):
                    filepath = os.path.join(row['ImageFolder'], filename)
                    try:
                        os.remove(filepath)
                    except Exception as e:
                        logging.warning(f"Could not delete file {filepath}: {str(e)}")
                try:
                    os.rmdir(row['ImageFolder'])
                except Exception as e:
                    logging.warning(f"Could not delete folder {row['ImageFolder']}: {str(e)}")
                    
            logging.info(f"Deleted Employee ID: {emp_id}")
            return jsonify({'success': True, 'message': 'Employee deleted successfully.'})
        except Exception as e:
            logging.error(f"Error deleting employee: {str(e)}")
            return jsonify({'success': False, 'message': f'Error deleting employee: {str(e)}'})

    @app.route('/api/reports/data')
    def api_reports_data():
        """Queries and returns attendance logs based on filters."""
        emp_search = request.args.get('search', '').strip()
        dept_filter = request.args.get('department', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        results = get_attendance_report_data(emp_search, dept_filter, start_date, end_date)
        return jsonify(results)

    @app.route('/api/reports/export/csv')
    def api_reports_export_csv():
        """Generates and triggers download of CSV attendance report."""
        emp_search = request.args.get('search', '').strip()
        dept_filter = request.args.get('department', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        results = get_attendance_report_data(emp_search, dept_filter, start_date, end_date)
        
        df_data = []
        for r in results:
            df_data.append({
                'Employee ID': r['employee_id'],
                'Name': r['name'],
                'Department': r['department'],
                'Date': r['date'],
                'Time': r['time'],
                'Status': r['status']
            })
            
        df = pd.DataFrame(df_data)
        
        csv_path = os.path.join("attendance", "attendance.csv")
        df.to_csv(csv_path, index=False)
        
        return send_file(csv_path, mimetype='text/csv', as_attachment=True, download_name="attendance_report.csv")

    @app.route('/api/reports/export/excel')
    def api_reports_export_excel():
        """Generates and returns styled Excel attendance sheets."""
        emp_search = request.args.get('search', '').strip()
        dept_filter = request.args.get('department', '').strip()
        start_date = request.args.get('start_date', '').strip()
        end_date = request.args.get('end_date', '').strip()
        
        results = get_attendance_report_data(emp_search, dept_filter, start_date, end_date)
        
        df_data = []
        for r in results:
            df_data.append({
                'Employee ID': r['employee_id'],
                'Employee Name': r['name'],
                'Department': r['department'],
                'Phone Number': r['phone'],
                'Email': r['email'],
                'Date': r['date'],
                'Time': r['time'],
                'Status': r['status']
            })
            
        df = pd.DataFrame(df_data)
        
        # Target filename: e.g. July_2026.xlsx based on current date
        month_year = datetime.now().strftime("%B_%Y")
        excel_filename = f"{month_year}.xlsx"
        excel_path = os.path.join("attendance", excel_filename)
        
        try:
            # Generate with pandas + openpyxl styling
            writer = pd.ExcelWriter(excel_path, engine='openpyxl')
            df.to_excel(writer, sheet_name="Attendance Logs", index=False, startrow=3)
            
            workbook = writer.book
            worksheet = writer.sheets["Attendance Logs"]
            
            # Add styled title headers
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Header banner
            worksheet.merge_cells("A1:H1")
            worksheet["A1"] = "CRT INDUSTRIES CHEMICAL LAB"
            worksheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            worksheet["A1"].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
            worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            
            worksheet.merge_cells("A2:H2")
            worksheet["A2"] = f"Attendance Authentication Report - Generated {datetime.now().strftime('%d-%m-%Y %H:%M')}"
            worksheet["A2"].font = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
            worksheet["A2"].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
            worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
            
            # Row heights
            worksheet.row_dimensions[1].height = 35
            worksheet.row_dimensions[2].height = 20
            worksheet.row_dimensions[4].height = 25 # Table columns header
            
            # Column headers style
            header_fill = PatternFill(start_color="0B0F19", end_color="0B0F19", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for col_num in range(1, 9):
                cell = worksheet.cell(row=4, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                
            # Data grid styling & width auto-fit
            for row in range(5, worksheet.max_row + 1):
                worksheet.row_dimensions[row].height = 20
                for col in range(1, 9):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.font = Font(name="Calibri", size=11)
                    
                    # Check status highlights
                    if col == 8: # Status col
                        if cell.value == "Present":
                            cell.font = Font(name="Calibri", size=11, bold=True, color="2E7D32")
                        else:
                            cell.font = Font(name="Calibri", size=11, bold=True, color="C62828")
            
            from openpyxl.utils import get_column_letter
            # Autofit column widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    # Ignore first two merged banner rows for width calculation
                    if cell.row in (1, 2):
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            writer.close()
            logging.info(f"Excel report saved successfully to {excel_path}.")
            return send_file(excel_path, as_attachment=True, download_name=excel_filename)
            
        except Exception as e:
            logging.error(f"Excel export failed: {str(e)}")
            # Send raw excel file as fallback
            df.to_excel(excel_path, index=False)
            return send_file(excel_path, as_attachment=True, download_name=excel_filename)

    @app.route('/api/departments')
    def api_departments():
        """Returns a list of unique departments currently registered in the system."""
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT Department FROM employees WHERE Department IS NOT NULL AND Department != ''")
        depts = [r['Department'] for r in cursor.fetchall()]
        conn.close()
        return jsonify(depts)

    @app.route('/dataset/<path:filename>')
    def serve_dataset(filename):
        if not session.get('logged_in'):
            return Response("Unauthorized", status=401)
        return send_from_directory('dataset', filename)

    @app.route('/api/test_verify')
    def api_test_verify():
        camera.latest_match = {
            'employee_id': 'hari456',
            'name': 'harihharan',
            'department': 'Chemical Analysis',
            'time': datetime.now().strftime("%H:%M:%S"),
            'date': datetime.now().strftime("%d-%m-%Y"),
            'status': 'Present',
            'timestamp': time.time()
        }
        return jsonify({"success": True, "message": "Verification simulation triggered."})

    @app.route('/settings')
    def settings_page():
        camera.system_mode = 'passive'
        return render_template('settings.html')

    @app.route('/help')
    def help_page():
        camera.system_mode = 'passive'
        return render_template('help.html')
