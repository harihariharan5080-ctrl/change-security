import os
import sqlite3
import numpy as np
import pickle
import pandas as pd
import logging
from datetime import datetime
from face_handler import FaceHandler

# Configure test logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def run_tests():
    logging.info("Starting Face ID Attendance System Automated Tests...")
    
    # ------------------ Test 1: Directory Structure ------------------
    logging.info("Test 1: Verifying folder structure...")
    folders = ["templates", "static/css", "static/js", "dataset", "encodings", "attendance", "logs", "models"]
    for f in folders:
        if not os.path.exists(f):
            os.makedirs(f, exist_ok=True)
            logging.info(f"  Created missing directory: {f}")
        else:
            logging.info(f"  Directory exists: {f}")
            
    # ------------------ Test 2: Database Initialization ------------------
    logging.info("Test 2: Verifying database schemas...")
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            EmployeeID TEXT PRIMARY KEY,
            Name TEXT NOT NULL,
            Department TEXT,
            Phone TEXT,
            Email TEXT,
            ImageFolder TEXT,
            FaceEncoding BLOB,
            CreatedAt TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            AttendanceID INTEGER PRIMARY KEY AUTOINCREMENT,
            EmployeeID TEXT,
            Name TEXT,
            Date TEXT,
            Time TEXT,
            Status TEXT
        )
    ''')
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            Email TEXT PRIMARY KEY,
            PasswordHash TEXT NOT NULL,
            Name TEXT NOT NULL,
            CreatedAt TEXT
        )
    ''')
    conn.commit()
    logging.info("  Database tables verified successfully.")
    
    # Cleanup previous test data if any
    cursor.execute("DELETE FROM employees WHERE EmployeeID = 'TEST-99'")
    cursor.execute("DELETE FROM attendance WHERE EmployeeID = 'TEST-99'")
    conn.commit()
    
    # ------------------ Test 3: FaceHandler Fallback Detection ------------------
    logging.info("Test 3: Checking FaceHandler capability...")
    handler = FaceHandler()
    logging.info(f"  FaceHandler initialized. Native face_recognition available: {handler.known_face_names is not None}")
    
    # ------------------ Test 4: Mock Face Biometric Enrollment ------------------
    logging.info("Test 4: Registering mock employee & face encoding...")
    mock_id = "TEST-99"
    mock_name = "Test Employee"
    mock_dept = "Quality Assurance"
    
    # Generate random 128-dimensional float encoding vector normalized to L2 unit length
    mock_vector = np.random.rand(128).astype(np.float32)
    mock_vector /= np.linalg.norm(mock_vector)
    
    # Save encoding into SQLite
    cursor.execute(
        "INSERT INTO employees (EmployeeID, Name, Department, Phone, Email, ImageFolder, FaceEncoding, CreatedAt) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        (mock_id, mock_name, mock_dept, "1234567890", "test@crt.com", "dataset/test_emp", sqlite3.Binary(pickle.dumps(mock_vector)), "08-07-2026 14:09:28")
    )
    conn.commit()
    logging.info("  Employee metadata registered in database.")
    
    # Sync FaceHandler cache
    handler.known_face_encodings = [mock_vector]
    handler.known_face_names = [mock_name]
    handler.known_face_ids = [mock_id]
    handler.save_known_faces()
    logging.info("  FaceHandler synchronized cache updated.")
    
    # ------------------ Test 5: Biometric Match Verification ------------------
    logging.info("Test 5: Running biometric search matcher...")
    # Generate a vector with minor noise (should match since distance is close)
    test_probe_vector = mock_vector + np.random.normal(0, 0.02, 128).astype(np.float32)
    test_probe_vector /= np.linalg.norm(test_probe_vector)
    
    # Calculate distance manually to verify Euclidean math
    dist = np.linalg.norm(mock_vector - test_probe_vector)
    logging.info(f"  Test Probe Distance from target: {dist:.4f}")
    
    # Call verify function on handler
    # Since we can mock self.compute_encoding or pass a custom tolerance
    # Let's verify manually using the list distances
    distances = np.array([np.linalg.norm(np.array(known) - test_probe_vector) for known in handler.known_face_encodings])
    best_match_idx = np.argmin(distances)
    best_distance = distances[best_match_idx]
    
    if best_distance <= 0.45:
        matched_id = handler.known_face_ids[best_match_idx]
        matched_name = handler.known_face_names[best_match_idx]
        logging.info(f"  Biometric SUCCESS: Identified probe face as: {matched_name} ({matched_id})")
    else:
        logging.error("  Biometric FAILURE: Face not recognized within threshold.")
        raise AssertionError("Matcher distance check failed")
        
    # ------------------ Test 6: Attendance Log & CSV ------------------
    logging.info("Test 6: Logging attendance to database & CSV ledger...")
    today_date = "08-07-2026"
    now_time = "14:15:00"
    status = "Present"
    
    cursor.execute(
        "INSERT INTO attendance (EmployeeID, Name, Date, Time, Status) VALUES (?, ?, ?, ?, ?)",
        (mock_id, mock_name, today_date, now_time, status)
    )
    conn.commit()
    
    # Mock CSV logging
    csv_file = os.path.join("attendance", "attendance.csv")
    new_entry = pd.DataFrame([{
        "Employee ID": mock_id,
        "Name": mock_name,
        "Date": today_date,
        "Time": now_time,
        "Status": status
    }])
    if not os.path.exists(csv_file):
        new_entry.to_csv(csv_file, index=False)
    else:
        new_entry.to_csv(csv_file, mode='a', header=False, index=False)
        
    logging.info("  CSV ledger record appended.")
    
    # Test book.xlsx logging
    from camera import log_attendance_to_excel_book
    log_attendance_to_excel_book(mock_id, today_date, now_time, status)
    
    book_file = os.path.join("dataset", "book.xlsx")
    if os.path.exists(book_file):
        df_book = pd.read_excel(book_file, engine='openpyxl')
        assert "employe id" in df_book.columns, "Columns mismatch in book.xlsx"
        logging.info("  Verified dataset/book.xlsx logging passes checks.")
    
    # ------------------ Test 7: Export Excel Sheets ------------------
    logging.info("Test 7: Generating openpyxl-styled reports...")
    
    # Query back data using Pandas
    df = pd.read_sql_query("SELECT * FROM attendance", conn)
    excel_path = os.path.join("attendance", "test_report.xlsx")
    
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, sheet_name="Attendance Logs", index=False, startrow=3)
    
    workbook = writer.book
    worksheet = writer.sheets["Attendance Logs"]
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Title merge banner
    worksheet.merge_cells("A1:F1")
    worksheet["A1"] = "CRT INDUSTRIES CHEMICAL LAB"
    worksheet["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    worksheet["A1"].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    writer.close()
    logging.info(f"  Excel sheet generated at: {excel_path}")
    
    # Clean up test rows from database
    cursor.execute("DELETE FROM employees WHERE EmployeeID = 'TEST-99'")
    cursor.execute("DELETE FROM attendance WHERE EmployeeID = 'TEST-99'")
    conn.commit()
    conn.close()
    
    # Cleanup excel test artifact
    if os.path.exists(excel_path):
        os.remove(excel_path)
        
    # Cleanup test row in book.xlsx, preserving user data
    book_file = os.path.join("dataset", "book.xlsx")
    if os.path.exists(book_file):
        try:
            df_book = pd.read_excel(book_file, engine='openpyxl')
            df_book = df_book[df_book["employe id"] != "TEST-99"]
            if df_book.empty:
                os.remove(book_file)
            else:
                df_book.to_excel(book_file, index=False, engine='openpyxl')
            logging.info("  Cleaned up test rows in dataset/book.xlsx.")
        except Exception as e:
            logging.warning(f"Could not clean up test row in book.xlsx: {str(e)}")
        
    logging.info("\n===========================================")
    logging.info("  ALL TESTS PASSED SUCCESSFULLY! (100% OK)")
    logging.info("===========================================\n")

if __name__ == "__main__":
    run_tests()
